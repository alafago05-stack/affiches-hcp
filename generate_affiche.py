#!/usr/bin/env python3
"""
generate_affiche.py (v3) — Génère des affiches HCP multilingues (FR / AR / EN)
à partir d'un fichier Excel d'indicateurs, avec une mise en page riche calquée
sur le modèle fourni :

  1. Introduction + lexique (icônes)
  2. Sous-utilisation de la main-d'œuvre (barres de progression)
  3. Répartition de la population en âge de travailler (arbre / boîtes imbriquées)
  4. Tableau comparatif Urbain / Rural / Ensemble
  5. Participation au marché du travail (3 jauges circulaires)
  6. Points clés (encadrés numérotés)

L'arabe est rendu en RTL (reshaping + bidi) avec mise en page miroir
(logo/titre inversés, texte aligné à droite, barres qui se remplissent
de droite à gauche, ordre des colonnes de tableau inversé, etc.)

Usage :
    python generate_affiche.py data.xlsx --list
    python generate_affiche.py data.xlsx --sheet 2025 --demo --lang fr --output affiche_fr.png
    python generate_affiche.py data.xlsx --sheet 2025 --demo --lang ar --output affiche_ar.png
    python generate_affiche.py data.xlsx --sheet 2025 --demo --lang en --output affiche_en.png
    python generate_affiche.py data.xlsx --sheet 2025 --config ma_config.json --lang fr --output out.png
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageFilter

import matplotlib

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_ARABIC_FALLBACK_LIBS = True
except ImportError:
    HAS_ARABIC_FALLBACK_LIBS = False

matplotlib.use("Agg")
import matplotlib.pyplot as plt

HERE = Path(__file__).parent

# ==========================================================================
# 1. PARSING EXCEL (format HCP "ENE-Indicateurs désagrégés")
# ==========================================================================

def parse_ene_excel(path: str, sheet: str) -> dict:
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet}' introuvable. Feuilles dispo: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    blocks, i, n = {}, 0, len(rows)
    while i < n:
        row = rows[i]
        title = row[0] if row else None
        if title and i + 1 < n and rows[i + 1][1]:
            milieu_row = rows[i + 1]
            milieux, col_idx = [], []
            for c in range(1, len(milieu_row), 2):
                if milieu_row[c]:
                    milieux.append(milieu_row[c])
                    col_idx.append(c)
            i += 3
            data_rows = {}
            while i < n and rows[i] and rows[i][0] and not str(rows[i][0]).startswith(
                ("Champ", "Référence", "Unité", "Source")
            ):
                r = rows[i]
                vals = {}
                for milieu, c in zip(milieux, col_idx):
                    valeur = r[c] if c < len(r) else None
                    signe = r[c + 1] if c + 1 < len(r) else None
                    vals[milieu] = {"valeur": valeur, "signe": signe}
                data_rows[r[0]] = vals
                i += 1
            meta = {"champ": "", "reference": "", "unite": "", "source": ""}
            while i < n and rows[i] and rows[i][0] and str(rows[i][0]).startswith(
                ("Champ", "Référence", "Unité", "Source")
            ):
                text = str(rows[i][0])
                key, _, val = text.partition(":")
                key = key.strip().lower()
                if key.startswith("champ"):
                    meta["champ"] = val.strip()
                elif key.startswith("réf"):
                    meta["reference"] = val.strip()
                elif key.startswith("unit"):
                    meta["unite"] = val.strip()
                elif key.startswith("source"):
                    meta["source"] = val.strip()
                i += 1
            blocks[title] = {"milieux": milieux, "rows": data_rows, **meta}
        else:
            i += 1
    return blocks


def get_value(blocks, titre, ligne="Ensemble", milieu="Ensemble"):
    if titre not in blocks:
        raise KeyError(f"Tableau introuvable: {titre!r}. Utilisez --list.")
    bloc = blocks[titre]
    if ligne not in bloc["rows"]:
        raise KeyError(f"Ligne {ligne!r} introuvable dans {titre!r}. Dispo: {list(bloc['rows'])}")
    if milieu not in bloc["rows"][ligne]:
        raise KeyError(f"Milieu {milieu!r} introuvable. Dispo: {bloc['milieux']}")
    return bloc["rows"][ligne][milieu]["valeur"], bloc["unite"]


def format_number(valeur, unite, lang="fr"):
    if valeur is None:
        return {"fr": "n.d.", "en": "n/a", "ar": "غ.م"}[lang]
    sep = "," if lang != "en" else "."
    if "%" in (unite or ""):
        return f"{valeur:.1f}".replace(".", sep) + "%"
    if "millier" in (unite or "").lower():
        return f"{int(round(valeur * 1000)):,}".replace(",", " ")
    if isinstance(valeur, float) and valeur.is_integer():
        return f"{int(valeur):,}".replace(",", " ")
    return str(valeur)


def pct_value(valeur, unite):
    if valeur is None or "%" not in (unite or ""):
        return None
    return float(valeur)


def hex_to_rgb(hex_color):
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


FALLBACK_KEYWORDS = {
    "TA": ["taux d'activité"],
    "TE": ["taux d'emploi"],
    "TC": ["taux de chômage"],
    "TS": ["taux de sous-emploi", "sous-emploi"],
}


def _resolve_title(blocks, code):
    """Le libellé exact des tableaux varie parfois d'une feuille/année à
    l'autre dans les fichiers HCP (ex: 'Le taux d'activité (%)' en 2019-2020
    vs 'Le taux d'activité par sexe et par milieu' en 2022-2025). On essaie
    d'abord le titre exact connu, puis on cherche par mot-clé."""
    titre = INDICATOR_TITLES[code]
    if titre in blocks:
        return titre
    for kw in FALLBACK_KEYWORDS.get(code, []):
        for t in blocks:
            if kw in t.lower():
                return t
    raise KeyError(
        f"Aucun tableau ne correspond à l'indicateur {code!r} dans cette feuille. "
        f"Utilisez --list sur ce fichier/feuille pour voir les titres disponibles."
    )


def get_rate(blocks, code, milieu="Ensemble", ligne="Ensemble", lang="fr"):
    """Raccourci pour récupérer un indicateur (TA/TE/TC/TS) déjà formaté +
    sa valeur en % brute. Utilisé par les 4 types d'affiche."""
    titre = _resolve_title(blocks, code)
    valeur, unite = get_value(blocks, titre, ligne, milieu)
    return format_number(valeur, unite, lang), pct_value(valeur, unite)


# ==========================================================================
# 2. TRADUCTIONS (chrome de l'interface)
# ==========================================================================

STRINGS = {
    "fr": {
        "intro_title": "Introduction",
        "lexique_title": "Lexique",
        "underuse_title": "Sous-utilisation de la main-d'œuvre",
        "breakdown_title": "Répartition de la population en âge de travailler",
        "table_title": "Comparaison Urbain / Rural / Ensemble",
        "participation_title": "Participation au marché du travail",
        "notes_title": "Points clés",
        "col_indicator": "Indicateur",
        "col_urban": "Urbain",
        "col_rural": "Rural",
        "col_ensemble": "Ensemble",
        "working_age": "Population en âge de travailler",
        "active": "Population active",
        "inactive": "Hors force de travail",
        "source_label": "Source",
        "age_chart_title": "Taux de chômage selon l'âge",
        "sector_title": "Répartition des travailleurs par secteur d'activité",
        "conditions_title": "Conditions d'emploi",
        "diploma_title": "Chômage et diplôme",
    },
    "en": {
        "intro_title": "Introduction",
        "lexique_title": "Glossary",
        "underuse_title": "Labour underutilization",
        "breakdown_title": "Breakdown of the working-age population",
        "table_title": "Urban / Rural / Total comparison",
        "participation_title": "Labour market participation",
        "notes_title": "Key takeaways",
        "col_indicator": "Indicator",
        "col_urban": "Urban",
        "col_rural": "Rural",
        "col_ensemble": "Total",
        "working_age": "Working-age population",
        "active": "Labour force",
        "inactive": "Outside the labour force",
        "source_label": "Source",
        "age_chart_title": "Unemployment rate by age group",
        "sector_title": "Workers by sector of activity",
        "conditions_title": "Working conditions",
        "diploma_title": "Unemployment and education level",
    },
    "ar": {
        "intro_title": "مقدمة",
        "lexique_title": "المعجم",
        "underuse_title": "الاستخدام غير الكامل لليد العاملة",
        "breakdown_title": "توزيع الساكنة في سن النشاط",
        "table_title": "مقارنة حضري / قروي / المجموع",
        "participation_title": "المشاركة في سوق الشغل",
        "notes_title": "أهم النقاط",
        "col_indicator": "المؤشر",
        "col_urban": "حضري",
        "col_rural": "قروي",
        "col_ensemble": "المجموع",
        "working_age": "السكان في سن النشاط",
        "active": "القوة العاملة",
        "inactive": "خارج القوى العاملة",
        "source_label": "المصدر",
        "age_chart_title": "معدل البطالة حسب الفئة العمرية",
        "sector_title": "توزيع الشغيلة حسب قطاع النشاط",
        "conditions_title": "ظروف الشغل",
        "diploma_title": "البطالة والمستوى الدراسي",
    },
}

# Contenu de démonstration traduit dans les 3 langues
DEMO_I18N = {
    "fr": {
        "kicker": "Enquête Nationale sur l'Emploi (ENE 2025)",
        "title": "Marché du travail dans la région Souss-Massa — année 2025",
        "subtitle": "Haut-Commissariat au Plan — Direction régionale de Souss-Massa",
        "rate_activity": "Taux d'activité",
        "rate_activity_sub": "Force de travail parmi la population en âge d'activité",
        "rate_employment": "Taux d'emploi",
        "rate_employment_sub": "Personnes en emploi parmi la population en âge d'activité",
        "rate_unemployment": "Taux de chômage",
        "rate_unemployment_sub": "Chômeurs parmi la force de travail",
        "intro_title": "Introduction",
        "intro": "Cette note présente les principaux résultats de l'Enquête Nationale sur l'Emploi pour la région Souss-Massa au titre de l'année 2025 : activité, emploi et chômage. Toutes les données portent sur l'ensemble de la population (milieux urbain et rural confondus).",
        "glossary_title": "Lexique",
        "glossary": [
            ("Emploi contre revenu", "Travail effectué contre un salaire ou un profit."),
            ("Taux d'activité", "Part de la force de travail dans la population en âge d'activité."),
            ("Taux d'emploi", "Part des personnes en emploi dans la population en âge d'activité."),
            ("Taux de chômage", "Part des chômeurs dans la force de travail."),
            ("Sous-emploi", "Actifs occupés souhaitant et disponibles à travailler davantage."),
        ],
        "dist_title": "Répartition de la population en âge d'activité",
        "total_pop_label": "Personnes en âge d'activité",
        "lf_label": "Force de travail",
        "out_label": "Hors force de travail",
        "lf_cards": ["Actifs occupés", "Chômeurs"],
        "out_cards": ["Inactifs"],
        "indic_title": "Indicateurs complémentaires (ensemble)",
        "indicators": ["Taux d'activité des femmes", "Taux de sous-emploi", "Taux de chômage des femmes", "Taux d'emploi des femmes"],
        "key_title": "Points saillants",
        "key_points": [
            "Une participation limitée au marché du travail, en particulier chez les femmes.",
            "Un taux de chômage qui reste élevé à l'échelle régionale.",
            "Un sous-emploi qui demeure présent parmi les actifs occupés.",
            "Des résultats qui orientent les politiques publiques régionales.",
        ],
        "source": "Source : Haut-Commissariat au Plan — Enquête Nationale sur l'Emploi (ENE), région Souss-Massa, 2025.",
    },
    "en": {
        "kicker": "National Employment Survey (ENE 2025)",
        "title": "Labour market in the Souss-Massa region — year 2025",
        "subtitle": "High Commission for Planning — Souss-Massa Regional Directorate",
        "rate_activity": "Activity rate",
        "rate_activity_sub": "Labour force out of the working-age population",
        "rate_employment": "Employment rate",
        "rate_employment_sub": "Employed persons out of the working-age population",
        "rate_unemployment": "Unemployment rate",
        "rate_unemployment_sub": "Unemployed out of the labour force",
        "intro_title": "Introduction",
        "intro": "This brief presents the key results of the National Employment Survey for the Souss-Massa region for the year 2025: activity, employment and unemployment. All figures refer to the total population (urban and rural combined).",
        "glossary_title": "Glossary",
        "glossary": [
            ("Employment for pay", "Work performed in exchange for pay or profit."),
            ("Activity rate", "Share of the labour force in the working-age population."),
            ("Employment rate", "Share of employed persons in the working-age population."),
            ("Unemployment rate", "Share of the unemployed in the labour force."),
            ("Underemployment", "Employed persons willing and available to work more."),
        ],
        "dist_title": "How is the working-age population distributed?",
        "total_pop_label": "Persons of working age",
        "lf_label": "Labour force",
        "out_label": "Outside the labour force",
        "lf_cards": ["Employed", "Unemployed"],
        "out_cards": ["Inactive"],
        "indic_title": "Additional indicators (total)",
        "indicators": ["Female activity rate", "Underemployment rate", "Female unemployment rate", "Female employment rate"],
        "key_title": "Highlights",
        "key_points": [
            "Limited labour market participation, especially among women.",
            "An unemployment rate that remains high at the regional level.",
            "Underemployment that remains present among the employed.",
            "Results that guide regional public policies.",
        ],
        "source": "Source: High Commission for Planning — National Employment Survey (ENE), Souss-Massa region, 2025.",
    },
    "ar": {
        "kicker": "البحث الوطني حول التشغيل (ENE 2025)",
        "title": "وضعية سوق الشغل بجهة سوس ماسة — سنة 2025",
        "subtitle": "المندوبية السامية للتخطيط — المديرية الجهوية لسوس ماسة",
        "rate_activity": "معدل النشاط",
        "rate_activity_sub": "القوى العاملة من الساكنة في سن النشاط",
        "rate_employment": "معدل الشغل",
        "rate_employment_sub": "المشتغلون من الساكنة في سن النشاط",
        "rate_unemployment": "معدل البطالة",
        "rate_unemployment_sub": "العاطلون من القوى العاملة",
        "intro_title": "المقدمة",
        "intro": "تقدم هذه المذكرة أبرز نتائج البحث الوطني حول التشغيل بجهة سوس ماسة برسم سنة 2025: النشاط، التشغيل والبطالة. جميع المعطيات تخص مجموع الساكنة (الوسطين الحضري والقروي معا).",
        "glossary_title": "المعجم",
        "glossary": [
            ("الشغل مقابل دخل", "العمل المنجز مقابل أجر أو ربح."),
            ("معدل النشاط", "نسبة القوى العاملة من الساكنة في سن النشاط."),
            ("معدل الشغل", "نسبة المشتغلين من الساكنة في سن النشاط."),
            ("معدل البطالة", "نسبة العاطلين من القوى العاملة."),
            ("الشغل الناقص", "المشتغلون الراغبون في العمل أكثر والمتوفرون لذلك."),
        ],
        "dist_title": "كيف تتوزع الساكنة في سن النشاط؟",
        "total_pop_label": "الأشخاص في سن النشاط",
        "lf_label": "القوة العاملة",
        "out_label": "خارج القوة العاملة",
        "lf_cards": ["المشتغلون", "العاطلون"],
        "out_cards": ["غير النشيطين"],
        "indic_title": "مؤشرات إضافية (المجموع)",
        "indicators": ["معدل النشاط لدى النساء", "معدل الشغل الناقص", "معدل البطالة لدى النساء", "معدل الشغل لدى النساء"],
        "key_title": "نقاط بارزة",
        "key_points": [
            "مشاركة محدودة في سوق الشغل، خاصة بين النساء.",
            "معدل بطالة لا يزال مرتفعا على مستوى الجهة.",
            "الشغل الناقص لا يزال حاضرا في صفوف المشتغلين.",
            "نتائج تساهم في توجيه السياسات العمومية الجهوية.",
        ],
        "source": "المصدر: المندوبية السامية للتخطيط — البحث الوطني حول التشغيل (ENE)، جهة سوس ماسة، 2025.",
    },
}

INDICATOR_TITLES = {
    "TA": "Le taux d'activité par sexe et par milieu",
    "TE": "Le taux d'emploi par sexe et par milieu",
    "TC": "Le taux de chômage par sexe et par milieu",
    "TS": "Le taux de sous-emploi par sexe et par milieu",
}

# ==========================================================================
# 3. PALETTE / POLICES
# ==========================================================================

# ==========================================================================
# 3. PALETTE / POLICES — thème sombre "data poster" néon
# ==========================================================================

BG = (233, 230, 218)            # fond extérieur (#E9E6DA)
PAGE = (247, 245, 236)          # fond de la fiche (#F7F5EC)
INK = (22, 50, 63)              # marine foncé — en-têtes, boîtes pleines (#16323F)
CARD = (255, 255, 255)          # cartes blanches (#FFFFFF)
CARD_ALT = (247, 245, 236)      # variante claire (#F7F5EC, glossaire/indicateurs)
CARD_BORDER = (228, 224, 210)   # bordure fine des cartes (#E4E0D2)
TRACK = (228, 224, 210)

TEXT = (34, 48, 56)             # texte principal (#223038)
TEXT_BODY = (68, 82, 90)        # texte de paragraphe (#44525A)
MUTED = (90, 104, 112)          # texte secondaire (#5A6870)
MUTED_ON_DARK = (157, 179, 188) # texte secondaire sur fond marine (#9DB3BC)
LIGHT_ON_DARK = (199, 210, 216) # texte clair sur fond marine (#C7D2D8)
KEYPOINTS_TEXT = (213, 222, 226)  # texte des points saillants (#D5DEE2)
WHITE = (255, 255, 255)

GREEN = (163, 181, 32)          # accent principal — olive (#A3B520)
GREEN_DARK = (126, 143, 20)     # accent au survol / plus profond (#7E8F14)
GOLD = (232, 161, 60)           # or/ambre — anneau chômage (#E8A13C)
CORAL = (196, 74, 62)           # rouge — baisse (delta négatif)

NAVY_DARK = INK  # alias conservé pour compatibilité interne
GREY = MUTED     # alias conservé pour compatibilité interne
CREAM = PAGE     # alias conservé pour compatibilité interne

F_LATIN_REG = str(HERE / "fonts" / "DejaVuSans.ttf")
F_LATIN_BOLD = str(HERE / "fonts" / "DejaVuSans-Bold.ttf")
F_AR_REG = str(HERE / "fonts" / "NotoSansArabic-Regular.ttf")
F_AR_BOLD = str(HERE / "fonts" / "NotoSansArabic-Bold.ttf")


def _detect_raqm():
    """Teste si Pillow dispose du moteur Raqm (nécessaire pour lier les
    lettres arabes nativement). Sur certains PC Windows, Pillow est installé
    sans cette option : on doit alors utiliser une solution de repli."""
    try:
        img = Image.new("RGB", (10, 10))
        d = ImageDraw.Draw(img)
        f = ImageFont.truetype(F_LATIN_REG, 10, layout_engine=ImageFont.Layout.RAQM)
        d.textbbox((0, 0), "test", font=f, direction="ltr")
        return True
    except Exception:
        return False


HAS_RAQM = _detect_raqm()
if not HAS_RAQM:
    print(
        "[info] Le moteur Raqm n'est pas disponible dans votre installation de Pillow.\n"
        "       Utilisation d'une méthode de repli pour l'arabe (arabic-reshaper + python-bidi)."
        + ("" if HAS_ARABIC_FALLBACK_LIBS else
           "\n       ATTENTION : ces librairies ne sont pas installées. Faites :\n"
           "       pip install arabic-reshaper python-bidi")
    )


def font(lang, size, bold=False):
    """Charge la police. Utilise le moteur Raqm (HarfBuzz) s'il est
    disponible, pour lier correctement les lettres arabes ; sinon, se rabat
    sur le moteur de base (le texte arabe est alors pré-formé manuellement,
    voir shape())."""
    if lang == "ar":
        path = F_AR_BOLD if bold else F_AR_REG
    else:
        path = F_LATIN_BOLD if bold else F_LATIN_REG
    if HAS_RAQM:
        return ImageFont.truetype(path, size, layout_engine=ImageFont.Layout.RAQM)
    return ImageFont.truetype(path, size)


def _direction(lang):
    return "rtl" if lang == "ar" else "ltr"


def shape(text, lang):
    """Avec Raqm : ne fait rien (Raqm gère la liaison des lettres et le sens
    de lecture au moment du dessin). Sans Raqm : reshape + bidi manuels,
    indispensables pour un rendu arabe correct."""
    text = str(text) if text is not None else ""
    if lang == "ar" and not HAS_RAQM and HAS_ARABIC_FALLBACK_LIBS and text:
        return get_display(arabic_reshaper.reshape(text))
    return text


def text_w(draw, text, fnt, lang="fr"):
    if HAS_RAQM:
        b = draw.textbbox((0, 0), text, font=fnt, direction=_direction(lang))
    else:
        b = draw.textbbox((0, 0), text, font=fnt)
    return b[2] - b[0]


def wrap(draw, text, fnt, max_width, lang="fr"):
    """Découpe `text` (texte brut, non transformé) en lignes qui tiennent dans
    max_width. Les lignes retournées restent en texte brut : c'est draw_line()
    qui applique le reshape/bidi juste avant le dessin, une seule fois."""
    words, lines, cur = text.split(), [], ""
    for w in words:
        test = (cur + " " + w).strip()
        if text_w(draw, shape(test, lang), fnt, lang) <= max_width or not cur:
            cur = test
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def draw_line(draw, x, y, text, fnt, fill, lang, align="left"):
    """Dessine une ligne de texte à la position x selon align."""
    text = shape(text, lang)
    w = text_w(draw, text, fnt, lang)
    kwargs = {"direction": _direction(lang)} if HAS_RAQM else {}
    if align == "right":
        draw.text((x - w, y), text, font=fnt, fill=fill, **kwargs)
    elif align == "center":
        draw.text((x - w / 2, y), text, font=fnt, fill=fill, **kwargs)
    else:
        draw.text((x, y), text, font=fnt, fill=fill, **kwargs)


# ---- icônes ----

def icon_briefcase(draw, cx, cy, s, color):
    w, h = s, s * 0.7
    draw.rounded_rectangle([cx - w / 2, cy - h / 2, cx + w / 2, cy + h / 2], radius=s * 0.08, outline=color, width=3)
    draw.rounded_rectangle([cx - w * 0.22, cy - h / 2 - s * 0.18, cx + w * 0.22, cy - h / 2], radius=s * 0.05, outline=color, width=3)
    draw.line([cx - w / 2, cy, cx + w / 2, cy], fill=color, width=3)


def icon_person(draw, cx, cy, s, color):
    r = s * 0.22
    draw.ellipse([cx - r, cy - s * 0.45, cx + r, cy - s * 0.45 + 2 * r], outline=color, width=3)
    draw.arc([cx - s * 0.32, cy - s * 0.05, cx + s * 0.32, cy + s * 0.6], 200, 340, fill=color, width=3)


def icon_group(draw, cx, cy, s, color):
    icon_person(draw, cx - s * 0.22, cy, s * 0.75, color)
    icon_person(draw, cx + s * 0.22, cy, s * 0.75, color)


def icon_clock(draw, cx, cy, s, color):
    r = s * 0.4
    draw.ellipse([cx - r, cy - r, cx + r, cy + r], outline=color, width=3)
    draw.line([cx, cy, cx, cy - r * 0.6], fill=color, width=3)
    draw.line([cx, cy, cx + r * 0.4, cy], fill=color, width=3)


def icon_badge(draw, cx, cy, s, color):
    r = s * 0.32
    draw.ellipse([cx - r, cy - r * 1.3, cx + r, cy + r * 0.7], outline=color, width=3)
    draw.polygon([(cx - r * 0.5, cy + r * 0.3), (cx - r * 0.15, cy + r * 1.1), (cx, cy + r * 0.6),
                  (cx + r * 0.15, cy + r * 1.1), (cx + r * 0.5, cy + r * 0.3)], outline=color, width=2)


def icon_heart(draw, cx, cy, s, color):
    r = s * 0.22
    draw.ellipse([cx - 2 * r, cy - r * 0.7, cx, cy + r * 0.9], outline=color, width=3)
    draw.ellipse([cx, cy - r * 0.7, cx + 2 * r, cy + r * 0.9], outline=color, width=3)
    draw.polygon([(cx - 2 * r * 0.9, cy + r * 0.4), (cx, cy + r * 1.7), (cx + 2 * r * 0.9, cy + r * 0.4)], outline=color, width=3)


def icon_shield(draw, cx, cy, s, color):
    w, h = s * 0.5, s * 0.62
    draw.polygon([
        (cx - w, cy - h * 0.7), (cx, cy - h), (cx + w, cy - h * 0.7),
        (cx + w, cy + h * 0.2), (cx, cy + h), (cx - w, cy + h * 0.2),
    ], outline=color, width=3)
    draw.line([cx - w * 0.4, cy, cx - w * 0.1, cy + h * 0.3], fill=color, width=3)
    draw.line([cx - w * 0.1, cy + h * 0.3, cx + w * 0.4, cy - h * 0.3], fill=color, width=3)


def icon_document(draw, cx, cy, s, color):
    w, h = s * 0.42, s * 0.58
    draw.rounded_rectangle([cx - w, cy - h, cx + w, cy + h], radius=4, outline=color, width=3)
    for i in range(3):
        ly = cy - h * 0.35 + i * h * 0.4
        draw.line([cx - w * 0.55, ly, cx + w * 0.55, ly], fill=color, width=2)


ICONS = {"briefcase": icon_briefcase, "person": icon_person, "group": icon_group,
         "clock": icon_clock, "badge": icon_badge, "heart": icon_heart,
         "shield": icon_shield, "document": icon_document}


def make_donut(value_pct, color, size_px=260):
    fig, ax = plt.subplots(figsize=(size_px / 100, size_px / 100), dpi=100)
    remainder = max(0, 100 - value_pct)
    ax.pie([value_pct, remainder], colors=[color, "#E7EAD8"], startangle=90,
           counterclock=False, wedgeprops=dict(width=0.28, edgecolor="#FFFFFF", linewidth=1))
    ax.set_aspect("equal")
    fig.patch.set_alpha(0)
    fig.subplots_adjust(0, 0, 1, 1)
    fig.canvas.draw()
    buf = fig.canvas.buffer_rgba()
    img = Image.frombuffer("RGBA", fig.canvas.get_width_height(), buf, "raw", "RGBA", 0, 1).copy()
    plt.close(fig)
    return img


def make_multi_donut(values, colors, size_px=300):
    fig, ax = plt.subplots(figsize=(size_px / 100, size_px / 100), dpi=100)
    ax.pie(values, colors=colors, startangle=90, counterclock=False,
           wedgeprops=dict(width=0.35, edgecolor="#FFFFFF", linewidth=2))
    ax.set_aspect("equal")
    fig.patch.set_alpha(0)
    fig.subplots_adjust(0, 0, 1, 1)
    fig.canvas.draw()
    buf = fig.canvas.buffer_rgba()
    img = Image.frombuffer("RGBA", fig.canvas.get_width_height(), buf, "raw", "RGBA", 0, 1).copy()
    plt.close(fig)
    return img


def make_line_chart(years, values, color, width_px=1000, height_px=280, lang="fr"):
    """Courbe d'évolution annuelle (style référence : ligne + marqueurs +
    étiquettes de valeur à chaque point), habillée pour le thème sombre."""
    fig, ax = plt.subplots(figsize=(width_px / 100, height_px / 100), dpi=100)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")

    xs = list(range(len(years)))
    ax.plot(xs, values, color=color, linewidth=2.5, marker="o",
             markersize=7, markerfacecolor=color, markeredgecolor="#FFFFFF", markeredgewidth=1.5, zorder=3)

    vmin, vmax = min(values), max(values)
    span = max(vmax - vmin, 1)
    for x, v in zip(xs, values):
        offset = span * 0.10
        label = f"{v:.1f}".replace(".", "," if lang != "en" else ".")
        ax.annotate(label, (x, v), textcoords="offset points", xytext=(0, 12 if v >= (vmin + vmax) / 2 else -18),
                    ha="center", fontsize=11, color="#0D3047", fontweight="bold")

    ax.set_xticks(xs)
    ax.set_xticklabels([str(y) for y in years], color="#6E767E", fontsize=10)
    ax.tick_params(axis="x", length=0)
    ax.set_yticks([])
    for spine in ("top", "right", "left"):
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color("#C7CEB8")
    ax.margins(x=0.06, y=0.35)
    fig.subplots_adjust(left=0.02, right=0.98, top=0.88, bottom=0.18)

    fig.canvas.draw()
    buf = fig.canvas.buffer_rgba()
    img = Image.frombuffer("RGBA", fig.canvas.get_width_height(), buf, "raw", "RGBA", 0, 1).copy()
    plt.close(fig)
    return img


# ==========================================================================
# 4. CONSTRUCTION DE L'AFFICHE
# ==========================================================================

def create_affiche(spec: dict, lang: str, output_path: str):
    S = STRINGS[lang]
    rtl = lang == "ar"
    align = "right" if rtl else "left"
    align_op = "left" if rtl else "right"
    W = 1240
    margin = 40
    content_w = W - 2 * margin
    img = Image.new("RGB", (W, 3200), PAGE)
    draw = ImageDraw.Draw(img)
    y = margin

    def card(x0, y0, x1, y1, fill=CARD, outline=CARD_BORDER, radius=14, **_):
        draw.rounded_rectangle([x0, y0, x1, y1], radius=radius, fill=fill, outline=outline, width=1)

    def badge(x, y0, num, on_accent=False, align="left"):
        """Petit badge carré numéroté, fidèle à la référence (pas de halo)."""
        size = 26
        x0 = x if align == "left" else x - size
        bg = GREEN if on_accent else INK
        fg = INK if on_accent else WHITE
        draw.rounded_rectangle([x0, y0, x0 + size, y0 + size], radius=8, fill=bg)
        f = font(lang, 13, bold=True)
        txt = shape(str(num), lang)
        w = text_w(draw, txt, f)
        draw.text((x0 + size / 2 - w / 2, y0 + size / 2 - 8), txt, font=f, fill=fg)
        return x0 + size + 10 if align == "left" else x0 - 10

    # ================= 1. EN-TÊTE =================
    header_h = 108
    card(margin, y, W - margin, y + header_h, fill=INK, outline=INK, radius=16)
    logo_path = spec.get("logo", str(HERE / "hcp_logo.png"))
    logo_box_w = 0
    if Path(logo_path).exists():
        logo = Image.open(logo_path).convert("RGBA")
        logo.thumbnail((110, 76))
        pad = 12
        plate_w, plate_h = logo.width + pad * 2, logo.height + pad * 2
        plate = Image.new("RGB", (plate_w, plate_h), WHITE)
        plate_img = Image.new("RGBA", (plate_w, plate_h), (255, 255, 255, 255))
        plate_img.paste(logo, (pad, pad), logo)
        logo_x = margin + 20 if rtl else W - margin - 20 - plate_w
        logo_y = y + (header_h - plate_h) / 2
        draw.rounded_rectangle([logo_x, logo_y, logo_x + plate_w, logo_y + plate_h], radius=10, fill=WHITE)
        img.paste(plate_img, (int(logo_x), int(logo_y)), plate_img)
        logo_box_w = plate_w + 20

    text_x0 = margin + 24 if not rtl else margin + 24 + logo_box_w
    text_x1 = W - margin - 24 - logo_box_w if not rtl else W - margin - 24
    text_area_w = text_x1 - text_x0
    tx = text_x0 if not rtl else text_x1
    ty = y + 18

    kicker = spec.get("kicker", "")
    if kicker:
        f_kicker = font(lang, 12, bold=True)
        kt = shape(kicker, lang)
        kw = text_w(draw, kt, f_kicker)
        pill_pad = 12
        px0 = tx if not rtl else tx - kw - 2 * pill_pad
        draw.rounded_rectangle([px0, ty, px0 + kw + 2 * pill_pad, ty + 24], radius=12, fill=GREEN)
        draw.text((px0 + pill_pad, ty + 4), kt, font=f_kicker, fill=INK)
        ty += 34

    f_title = font(lang, 22, bold=True)
    title_lines = wrap(draw, spec["title"], f_title, text_area_w, lang)
    for line in title_lines[:2]:
        draw_line(draw, tx, ty, line, f_title, WHITE, lang, align=align)
        ty += 28
    f_sub = font(lang, 13)
    for line in wrap(draw, spec.get("subtitle", ""), f_sub, text_area_w, lang)[:1]:
        draw_line(draw, tx, ty + 4, line, f_sub, MUTED_ON_DARK, lang, align=align)

    y += header_h + 18

    # ================= 2. TROIS TAUX ESSENTIELS (hero) =================
    hero = spec.get("hero_rates")
    if hero:
        hh = 128
        n = len(hero)
        gap = 16
        cw = (content_w - gap * (n - 1)) / n
        colors_ring = {"accent": GREEN, "ink": INK, "gold": GOLD}
        for idx, r in enumerate(hero):
            cx0 = margin + idx * (cw + gap)
            dark = r.get("style") == "dark"
            fillc = INK if dark else CARD
            card(cx0, y, cx0 + cw, y + hh, fill=fillc, outline=(INK if dark else CARD_BORDER), radius=14)
            ring_color = colors_ring.get(r.get("ring", "ink"), INK)
            dsize = 92
            donut_img = make_donut(r["pct"], "#%02X%02X%02X" % ring_color, size_px=dsize)
            # trou central de la couleur de la carte pour un effet anneau propre
            dx = cx0 + 20
            dy = y + (hh - dsize) / 2
            img.paste(donut_img, (int(dx), int(dy)), donut_img)
            f_val = font(lang, 20, bold=True)
            valcol = WHITE if dark else TEXT
            draw_line(draw, dx + dsize / 2, dy + dsize / 2 - 13, r["value"], f_val, valcol, lang, align="center")
            tx2 = dx + dsize + 16
            f_lbl = font(lang, 14, bold=True)
            lblcol = WHITE if dark else TEXT
            draw_line(draw, tx2, y + 30, r["label"], f_lbl, lblcol, lang, align="left")
            f_subl = font(lang, 11)
            subcol = MUTED_ON_DARK if dark else MUTED
            sub_lines = wrap(draw, r["sub"], f_subl, cw - (dx + dsize + 16 - cx0) - 14, lang)
            sy = y + 50
            for line in sub_lines[:3]:
                draw_line(draw, tx2, sy, line, f_subl, subcol, lang, align="left")
                sy += 15
        y += hh + 18

    # ================= 3. INTRODUCTION + LEXIQUE (côte à côte) =================
    intro = spec.get("intro")
    glossary = spec.get("glossary")
    if intro or glossary:
        gap = 16
        intro_w = content_w * (1 / 2.9)
        gloss_w = content_w - intro_w - gap
        if rtl:
            intro_x0, intro_x1 = W - margin - intro_w, W - margin
            gloss_x0, gloss_x1 = margin, margin + gloss_w
        else:
            intro_x0, intro_x1 = margin, margin + intro_w
            gloss_x0, gloss_x1 = W - margin - gloss_w, W - margin

        row_h = 190
        if intro:
            card(intro_x0, y, intro_x1, y + row_h)
            bx = badge(intro_x0 + 14 if not rtl else intro_x1 - 14, y + 14, intro.get("number", 1), align=align)
            f_h = font(lang, 15, bold=True)
            draw_line(draw, bx, y + 17, intro.get("title", S["intro_title"]), f_h, TEXT, lang, align=align)
            f_body = font(lang, 12)
            ix = intro_x0 + 16 if not rtl else intro_x1 - 16
            iy = y + 52
            for line in wrap(draw, intro["text"], f_body, intro_w - 32, lang):
                draw_line(draw, ix, iy, line, f_body, TEXT_BODY, lang, align=align)
                iy += 18.5
                if iy > y + row_h - 16:
                    break

        if glossary:
            card(gloss_x0, y, gloss_x1, y + row_h)
            bx = badge(gloss_x0 + 14 if not rtl else gloss_x1 - 14, y + 14, glossary.get("number", 2), align=align)
            f_h = font(lang, 15, bold=True)
            draw_line(draw, bx, y + 17, glossary.get("title", S["lexique_title"]), f_h, TEXT, lang, align=align)
            items = glossary["items"]
            n = len(items)
            gcell_gap = 8
            gcell_w = (gloss_w - 32 - gcell_gap * (n - 1)) / n
            gy0 = y + 52
            gcell_h = row_h - 52 - 14
            for idx, it in enumerate(items):
                order = idx if not rtl else (n - 1 - idx)
                gx0 = gloss_x0 + 16 + order * (gcell_w + gcell_gap)
                card(gx0, gy0, gx0 + gcell_w, gy0 + gcell_h, fill=CARD_ALT, outline=CARD_ALT, radius=10)
                draw.rounded_rectangle([gx0 + 10, gy0 + 10, gx0 + 32, gy0 + 13], radius=2, fill=GREEN)
                f_term = font(lang, 11, bold=True)
                ty2 = gy0 + 20
                for line in wrap(draw, it["term"], f_term, gcell_w - 20, lang)[:2]:
                    draw_line(draw, gx0 + 10, ty2, line, f_term, TEXT, lang, align="left")
                    ty2 += 14
                f_def = font(lang, 9)
                for line in wrap(draw, it["def"], f_def, gcell_w - 20, lang)[:5]:
                    draw_line(draw, gx0 + 10, ty2, line, f_def, MUTED, lang, align="left")
                    ty2 += 12.5
        y += row_h + 18

    # ================= 4. RÉPARTITION (diagramme de flux) =================
    dist = spec.get("distribution")
    if dist:
        box_h = 250
        card(margin, y, W - margin, y + box_h)
        bx = badge(margin + 16 if not rtl else W - margin - 16, y + 16, dist.get("number", 3), align=align)
        f_h = font(lang, 15, bold=True)
        draw_line(draw, bx, y + 19, dist.get("title", S["breakdown_title"]), f_h, TEXT, lang, align=align)

        total_col_w = 168
        if rtl:
            flow_x0, flow_x1 = margin + 16 + total_col_w + 12, W - margin - 16
            total_x0, total_x1 = margin + 16, margin + 16 + total_col_w
        else:
            flow_x0, flow_x1 = margin + 16, W - margin - 16 - total_col_w - 12
            total_x0, total_x1 = W - margin - 16 - total_col_w, W - margin - 16

        groups = dist["groups"]  # [{"cards":[...], "rate":..., "count":..., "label":...}, x2]
        gy0 = y + 56
        gh = (box_h - 56 - 16 - 12) / 2
        for gi, grp in enumerate(groups):
            gy = gy0 + gi * (gh + 12)
            cards_ = grp["cards"]
            nb_cards = len(cards_)
            big_w = 178
            if rtl:
                big_x1 = flow_x1
                big_x0 = big_x1 - big_w
                feed_x1 = big_x0 - 14
                feed_x0 = flow_x0
            else:
                big_x0 = flow_x0
                big_x1 = big_x0 + big_w
                feed_x0 = big_x1 + 14
                feed_x1 = flow_x1
            # boîte principale (sombre)
            card(big_x0, gy, big_x1, gy + gh, fill=INK, outline=INK, radius=10)
            f_bv = font(lang, 21, bold=True)
            draw_line(draw, (big_x0 + big_x1) / 2, gy + 10, grp["rate"], f_bv, WHITE, lang, align="center")
            f_bl = font(lang, 11, bold=True)
            draw_line(draw, (big_x0 + big_x1) / 2, gy + 38, grp["label"], f_bl, WHITE, lang, align="center")
            draw.line([(big_x0 + 16, gy + gh - 22), (big_x1 - 16, gy + gh - 22)], fill=(255, 255, 255, 60))
            f_bc = font(lang, 11)
            draw_line(draw, (big_x0 + big_x1) / 2, gy + gh - 18, grp["count"], f_bc, MUTED_ON_DARK, lang, align="center")

            # petites cartes nourricières
            fc_gap = 6
            fc_h = (gh - fc_gap * (nb_cards - 1)) / nb_cards
            for ci, c in enumerate(cards_):
                fy = gy + ci * (fc_h + fc_gap)
                card(feed_x0, fy, feed_x1, fy + fc_h, fill=INK, outline=INK, radius=8)
                cxm = (feed_x0 + feed_x1) / 2
                f_cv = font(lang, 13, bold=True)
                draw_line(draw, cxm, fy + 4, c["pct"], f_cv, WHITE, lang, align="center")
                f_cl = font(lang, 9)
                lines = wrap(draw, c["label"], f_cl, (feed_x1 - feed_x0) - 16, lang)
                cy2 = fy + 4 + 16
                for line in lines[:2]:
                    draw_line(draw, cxm, cy2, line, f_cl, LIGHT_ON_DARK, lang, align="center")
                    cy2 += 11
                f_cc = font(lang, 10, bold=True)
                draw_line(draw, cxm, fy + fc_h - 14, c["count"], f_cc, GREEN, lang, align="center")

        # colonne population totale
        tcx = (total_x0 + total_x1) / 2
        f_tl = font(lang, 12, bold=True)
        tl_lines = wrap(draw, dist["total_label"], f_tl, total_col_w - 10, lang)
        ty2 = y + box_h / 2 - 30
        for line in tl_lines[:2]:
            draw_line(draw, tcx, ty2, line, f_tl, TEXT, lang, align="center")
            ty2 += 16
        draw.line([tcx - 30, ty2 + 2, tcx + 30, ty2 + 2], fill=GREEN, width=3)
        f_tv = font(lang, 19, bold=True)
        draw_line(draw, tcx, ty2 + 10, dist["total_value"], f_tv, TEXT, lang, align="center")

        y += box_h + 18

    # ================= 5. INDICATEURS COMPLÉMENTAIRES =================
    indicators = spec.get("indicators")
    if indicators:
        items = indicators["items"]
        n = len(items)
        box_h = 96
        card(margin, y, W - margin, y + box_h)
        bar_x = margin + 16 if not rtl else W - margin - 16
        draw.rounded_rectangle([bar_x - (0 if not rtl else 4), y + 14, bar_x + (4 if not rtl else 0), y + 34],
                                 radius=2, fill=GREEN)
        f_h = font(lang, 15, bold=True)
        tx3 = bar_x + 12 if not rtl else bar_x - 12
        draw_line(draw, tx3, y + 16, indicators.get("title", ""), f_h, TEXT, lang, align=align)

        gap = 12
        cell_w = (content_w - 32 - gap * (n - 1)) / n
        cy0 = y + 46
        cell_h = box_h - 46 - 12
        for idx, it in enumerate(items):
            order = idx if not rtl else (n - 1 - idx)
            cx0 = margin + 16 + order * (cell_w + gap)
            card(cx0, cy0, cx0 + cell_w, cy0 + cell_h, fill=CARD_ALT, outline=CARD_ALT, radius=8)
            draw.rectangle([cx0, cy0, cx0 + cell_w, cy0 + 3], fill=GREEN)
            f_v = font(lang, 19, bold=True)
            draw_line(draw, cx0 + 12, cy0 + 12, it["value"], f_v, TEXT, lang, align="left")
            f_l = font(lang, 9, bold=True)
            ly = cy0 + 38
            for line in wrap(draw, it["label"], f_l, cell_w - 20, lang)[:2]:
                draw_line(draw, cx0 + 12, ly, line, f_l, MUTED, lang, align="left")
                ly += 12
        y += box_h + 18

    # ================= SECTIONS FLEXIBLES SUPPLÉMENTAIRES (comparatifs) =================
    compare = spec.get("compare_bars")
    if compare:
        section_title_y = y
        badge(margin if not rtl else W - margin - 26, y, compare.get("number", 6), align=align)
        f_h = font(lang, 16, bold=True)
        tx4 = margin + 36 if not rtl else W - margin - 36
        draw_line(draw, tx4, y + 4, compare["title"], f_h, TEXT, lang, align=align)
        y += 40
        entities = compare["entities"]
        items = compare["items"]
        n_entities = len(entities)
        colors = [GREEN, GOLD, (59, 111, 160), CORAL][:n_entities]
        row_h = 34 + 22 * n_entities
        box_h = row_h * len(items) + 16
        card(margin, y, W - margin, y + box_h, radius=14)
        by = y + 14
        f_ind = font(lang, 14, bold=True)
        f_ent = font(lang, 12)
        f_delta = font(lang, 12, bold=True)
        row_x0, row_x1 = margin + 20, W - margin - 20
        bar_area_w = (row_x1 - row_x0) - 110
        for it in items:
            draw_line(draw, row_x1 if rtl else row_x0, by, it["label"], f_ind, TEXT, lang, align=align)
            if n_entities == 2 and it["pcts"][0] is not None and it["pcts"][1] is not None:
                delta = it["pcts"][1] - it["pcts"][0]
                sign = "+" if delta >= 0 else ""
                dtxt = f"{sign}{delta:.1f} pt".replace(".", "," if lang != "en" else ".")
                dcolor = GREEN if delta >= 0 else CORAL
                draw_line(draw, row_x0 if rtl else row_x1, by, dtxt, f_delta, dcolor, lang,
                          align=("left" if rtl else "right"))
            ey = by + 22
            for idx, (ent_label, val_str, pct) in enumerate(zip(entities, it["values"], it["pcts"])):
                pct = pct or 0
                bar_full = int(bar_area_w * min(pct, 100) / 100)
                bar_y = ey + 9
                ecol = colors[idx % len(colors)]
                if rtl:
                    draw_line(draw, row_x1, ey, f"{ent_label} — {val_str}", f_ent, MUTED, lang, align="right")
                    draw.line([row_x1 - bar_area_w, bar_y, row_x1, bar_y], fill=TRACK, width=4)
                    draw.line([row_x1 - bar_full, bar_y, row_x1, bar_y], fill=ecol, width=4)
                else:
                    draw_line(draw, row_x0, ey, f"{ent_label} — {val_str}", f_ent, MUTED, lang, align="left")
                    draw.line([row_x0, bar_y, row_x0 + bar_area_w, bar_y], fill=TRACK, width=4)
                    draw.line([row_x0, bar_y, row_x0 + bar_full, bar_y], fill=ecol, width=4)
                ey += 22
            by += row_h
        y += box_h + 18

    trends = spec.get("trend_charts")
    if trends:
        badge(margin if not rtl else W - margin - 26, y, trends.get("number", 6), align=align)
        f_h = font(lang, 16, bold=True)
        tx4 = margin + 36 if not rtl else W - margin - 36
        draw_line(draw, tx4, y + 4, trends["title"], f_h, TEXT, lang, align=align)
        y += 40
        chart_w = content_w - 40
        chart_h = 190
        for chart in trends["charts"]:
            ch = chart_h + 40
            card(margin, y, W - margin, y + ch, radius=14)
            f_lbl = font(lang, 14, bold=True)
            label_x = W - margin - 18 if rtl else margin + 18
            draw_line(draw, label_x, y + 12, chart["label"], f_lbl, TEXT, lang, align=align)
            last_val = chart["values"][-1]
            last_txt = f"{last_val:.1f}".replace(".", "," if lang != "en" else ".") + "%"
            val_x = margin + 18 if rtl else W - margin - 18
            draw_line(draw, val_x, y + 12, last_txt, f_lbl, chart.get("color_rgb", GREEN), lang,
                      align=("left" if rtl else "right"))
            chart_img = make_line_chart(chart["years"], chart["values"], chart.get("color", "#A3B520"),
                                          width_px=chart_w, height_px=chart_h, lang=lang)
            img.paste(chart_img, (margin + 20, y + 34), chart_img)
            y += ch + 14
        y += 4

    # ================= 6. POINTS SAILLANTS =================
    key_points = spec.get("key_points")
    if key_points:
        items = key_points["items"]
        n = len(items)
        box_h = 108
        card(margin, y, W - margin, y + box_h, fill=INK, outline=INK, radius=14)
        bx = badge(margin + 16 if not rtl else W - margin - 16, y + 16, key_points.get("number", 5),
                    on_accent=True, align=align)
        f_h = font(lang, 15, bold=True)
        draw_line(draw, bx, y + 19, key_points.get("title", S["notes_title"]), f_h, WHITE, lang, align=align)

        gap = 14
        cell_w = (content_w - 32 - gap * (n - 1)) / n
        cy0 = y + 52
        for idx, kp in enumerate(items):
            order = idx if not rtl else (n - 1 - idx)
            cx0 = margin + 16 + order * (cell_w + gap)
            draw.rectangle([cx0, cy0, cx0 + cell_w, cy0 + 2], fill=GREEN)
            f_num = font(lang, 16, bold=True)
            draw_line(draw, cx0, cy0 + 8, kp["num"], f_num, GREEN, lang, align="left")
            f_txt = font(lang, 10)
            ty2 = cy0 + 28
            for line in wrap(draw, kp["text"], f_txt, cell_w - 6, lang)[:5]:
                draw_line(draw, cx0, ty2, line, f_txt, KEYPOINTS_TEXT, lang, align="left")
                ty2 += 13.5
        y += box_h + 14

    # ================= SOURCE =================
    src = spec.get("source", "")
    if src:
        f_src = font(lang, 11)
        draw_line(draw, W / 2, y, src, f_src, (122, 133, 127), lang, align="center")
        y += 20

    img = img.crop((0, 0, W, int(y + margin)))
    img.save(output_path)
    return output_path


# ==========================================================================
# 5. RÉSOLUTION DES VALEURS EXCEL DANS LE SPEC
# ==========================================================================

def build_from_excel(blocks, spec, lang):
    if isinstance(spec, dict):
        if "from_excel" in spec:
            fe = spec["from_excel"]
            valeur, unite = get_value(blocks, fe["titre"], fe.get("ligne", "Ensemble"), fe.get("milieu", "Ensemble"))
            out = {k: v for k, v in spec.items() if k != "from_excel"}
            out["valeur"] = format_number(valeur, unite, lang)
            p = pct_value(valeur, unite)
            if p is not None:
                out["pct"] = p
            return out
        return {k: build_from_excel(blocks, v, lang) for k, v in spec.items()}
    if isinstance(spec, list):
        return [build_from_excel(blocks, v, lang) for v in spec]
    return spec


def build_hero_rates(blocks, lang, accent_ring="accent"):
    """Les 3 taux essentiels (activité / emploi / chômage) — cartes en anneau,
    fidèles à la référence (1 carte foncée + 2 cartes claires)."""
    C = DEMO_I18N[lang]
    v_ta, p_ta = get_rate(blocks, "TA", lang=lang)
    v_te, p_te = get_rate(blocks, "TE", lang=lang)
    v_tc, p_tc = get_rate(blocks, "TC", lang=lang)
    return [
        {"value": v_ta, "pct": p_ta or 0, "label": C["rate_activity"], "sub": C["rate_activity_sub"],
         "style": "dark", "ring": "accent"},
        {"value": v_te, "pct": p_te or 0, "label": C["rate_employment"], "sub": C["rate_employment_sub"],
         "style": "light", "ring": "ink"},
        {"value": v_tc, "pct": p_tc or 0, "label": C["rate_unemployment"], "sub": C["rate_unemployment_sub"],
         "style": "light", "ring": "gold"},
    ]


def build_glossary(lang, number=2):
    C = DEMO_I18N[lang]
    return {"number": number, "title": C["glossary_title"],
            "items": [{"term": term, "def": d} for term, d in C["glossary"]]}


def build_distribution(blocks, lang, number=3):
    """Diagramme de répartition de la population en âge d'activité — construit
    à partir de vraies données Excel (table population en âge de travailler)."""
    C = DEMO_I18N[lang]
    titre_pop = "La population en âge de travailler par type d'activité (3) et par milieu"
    total_v, unite = get_value(blocks, titre_pop, "Ensemble", "Ensemble")
    actif_v, _ = get_value(blocks, titre_pop, "Actif occupé", "Ensemble")
    chomeur_v, _ = get_value(blocks, titre_pop, "Chômeur", "Ensemble")
    inactif_v, _ = get_value(blocks, titre_pop, "Inactif", "Ensemble")
    lf_v = actif_v + chomeur_v

    v_ta, p_ta = get_rate(blocks, "TA", lang=lang)
    p_out = 100 - (p_ta or 0)

    def fmt(x):
        return format_number(x, unite, lang)

    pct_actif = (actif_v / lf_v * 100) if lf_v else 0
    pct_chomeur = (chomeur_v / lf_v * 100) if lf_v else 0

    return {
        "number": number,
        "title": C["dist_title"],
        "total_label": C["total_pop_label"],
        "total_value": fmt(total_v),
        "groups": [
            {
                "label": C["lf_label"], "rate": v_ta, "count": fmt(lf_v),
                "cards": [
                    {"pct": format_number(pct_actif, "%", lang), "count": fmt(actif_v), "label": C["lf_cards"][0]},
                    {"pct": format_number(pct_chomeur, "%", lang), "count": fmt(chomeur_v), "label": C["lf_cards"][1]},
                ],
            },
            {
                "label": C["out_label"], "rate": format_number(p_out, "%", lang), "count": fmt(inactif_v),
                "cards": [
                    {"pct": format_number(100.0, "%", lang), "count": fmt(inactif_v), "label": C["out_cards"][0]},
                ],
            },
        ],
    }


def build_indicators(blocks, lang, number=4):
    C = DEMO_I18N[lang]
    v_ta_f, _ = get_rate(blocks, "TA", "Ensemble", "Féminin", lang=lang)
    v_ts, _ = get_rate(blocks, "TS", lang=lang)
    v_tc_f, _ = get_rate(blocks, "TC", "Ensemble", "Féminin", lang=lang)
    v_te_f, _ = get_rate(blocks, "TE", "Ensemble", "Féminin", lang=lang)
    values = [v_ta_f, v_ts, v_tc_f, v_te_f]
    return {"number": number, "title": C["indic_title"],
            "items": [{"value": v, "label": lbl} for v, lbl in zip(values, C["indicators"])]}


def build_key_points(lang, number=5):
    C = DEMO_I18N[lang]
    return {"number": number, "title": C["key_title"],
            "items": [{"num": f"0{i+1}", "text": txt} for i, txt in enumerate(C["key_points"])]}


def build_demo_spec(blocks, lang):
    C = DEMO_I18N[lang]
    return {
        "kicker": C["kicker"],
        "title": C["title"],
        "subtitle": C["subtitle"],
        "hero_rates": build_hero_rates(blocks, lang),
        "intro": {"number": 1, "title": C["intro_title"], "text": C["intro"]},
        "glossary": build_glossary(lang, number=2),
        "distribution": build_distribution(blocks, lang, number=3),
        "indicators": build_indicators(blocks, lang, number=4),
        "key_points": build_key_points(lang, number=5),
        "source": C["source"],
    }


# ==========================================================================
# 5bis. TYPES 2, 3, 4 — AFFICHES COMPARATIVES
# Réutilisent la même zone FIXE (hero_rates/intro/glossary/distribution) que
# le poster standard. Seule la section après "distribution" change : une
# section comparative (barres ou courbes) remplace "indicators", et
# key_points passe en position 5.
# ==========================================================================

COMPARE_I18N = {
    "fr": {
        "intro_year": "Cette fiche compare l'évolution du marché du travail régional entre {a} et {b}, à partir des résultats de l'Enquête Nationale sur l'Emploi.",
        "intro_region": "Cette fiche compare le marché du travail entre {a} et {b} pour l'année {year}, à partir des résultats de l'Enquête Nationale sur l'Emploi.",
        "intro_quarter": "Cette fiche compare l'évolution trimestrielle du marché du travail en {year}, à partir des résultats de l'Enquête Nationale sur l'Emploi.",
        "title_year": "Comparatif {a} — {b}",
        "title_region": "{a} — {b}",
        "title_quarter": "Évolution trimestrielle {year}",
        "subtitle_year": "Marché du travail — évolution annuelle",
        "subtitle_region": "Marché du travail — comparaison régionale ({year})",
        "subtitle_quarter": "Marché du travail par trimestre",
        "compare_section_title": "Indicateurs comparés",
        "indicators": ["Taux d'activité", "Taux d'emploi", "Taux de chômage"],
        "source": "Source : Haut-Commissariat au Plan — Enquête Nationale sur l'Emploi",
    },
    "en": {
        "intro_year": "This factsheet compares the regional labour market between {a} and {b}, based on the National Labour Force Survey.",
        "intro_region": "This factsheet compares the labour market between {a} and {b} for {year}, based on the National Labour Force Survey.",
        "intro_quarter": "This factsheet compares the quarterly evolution of the labour market in {year}, based on the National Labour Force Survey.",
        "title_year": "Comparison {a} — {b}",
        "title_region": "{a} — {b}",
        "title_quarter": "Quarterly evolution {year}",
        "subtitle_year": "Labour market — year-over-year",
        "subtitle_region": "Labour market — regional comparison ({year})",
        "subtitle_quarter": "Labour market by quarter",
        "compare_section_title": "Compared indicators",
        "indicators": ["Activity rate", "Employment rate", "Unemployment rate"],
        "source": "Source: Haut-Commissariat au Plan — National Labour Force Survey",
    },
    "ar": {
        "intro_year": "تقارن هذه المذكرة تطور سوق الشغل الجهوي بين {a} و {b}، اعتمادا على نتائج البحث الوطني حول التشغيل.",
        "intro_region": "تقارن هذه المذكرة سوق الشغل بين {a} و {b} برسم سنة {year}، اعتمادا على نتائج البحث الوطني حول التشغيل.",
        "intro_quarter": "تقارن هذه المذكرة التطور الفصلي لسوق الشغل خلال سنة {year}، اعتمادا على نتائج البحث الوطني حول التشغيل.",
        "title_year": "مقارنة {a} — {b}",
        "title_region": "{a} — {b}",
        "title_quarter": "التطور الفصلي {year}",
        "subtitle_year": "سوق الشغل — تطور سنوي",
        "subtitle_region": "سوق الشغل — مقارنة جهوية ({year})",
        "subtitle_quarter": "سوق الشغل حسب الفصل",
        "compare_section_title": "المؤشرات المقارنة",
        "indicators": ["معدل النشاط", "معدل الشغل", "معدل البطالة"],
        "source": "المصدر: المندوبية السامية للتخطيط — البحث الوطني حول التشغيل",
    },
}


def build_year_compare_spec(blocks_by_year, year_a, year_b, lang):
    T = COMPARE_I18N[lang]
    C = DEMO_I18N[lang]
    years_sorted = sorted(blocks_by_year.keys(), key=lambda y: int(y))
    chart_colors = {"TA": "#4E9A50", "TE": "#3B6FA0", "TC": "#C44A3E"}

    charts = []
    for code, label in zip(["TA", "TE", "TC"], T["indicators"]):
        vals, valid_years = [], []
        for yr in years_sorted:
            try:
                _, p = get_rate(blocks_by_year[yr], code, lang=lang)
            except KeyError:
                continue
            if p is not None:
                vals.append(p)
                valid_years.append(yr)
        if vals:
            charts.append({"label": label, "years": valid_years, "values": vals,
                            "color": chart_colors[code], "color_rgb": hex_to_rgb(chart_colors[code])})

    blocks_b = blocks_by_year[str(year_b)]

    return {
        "kicker": C["kicker"],
        "title": T["title_year"].format(a=year_a, b=year_b),
        "subtitle": T["subtitle_year"],
        "hero_rates": build_hero_rates(blocks_b, lang),
        "intro": {"number": 1, "title": C["intro_title"], "text": T["intro_year"].format(a=year_a, b=year_b)},
        "glossary": build_glossary(lang, number=2),
        "distribution": build_distribution(blocks_b, lang, number=3),
        "trend_charts": {"number": 4, "title": T["compare_section_title"], "charts": charts},
        "key_points": build_key_points(lang, number=5),
        "source": T["source"],
    }


def build_region_compare_spec(blocks_a, blocks_b, region_a, region_b, year_label, lang):
    T = COMPARE_I18N[lang]
    C = DEMO_I18N[lang]
    items = []
    for code, label in zip(["TA", "TE", "TC"], T["indicators"]):
        va, pa = get_rate(blocks_a, code, lang=lang)
        vb, pb = get_rate(blocks_b, code, lang=lang)
        items.append({"label": label, "values": [va, vb], "pcts": [pa, pb]})

    return {
        "kicker": C["kicker"],
        "title": T["title_region"].format(a=region_a, b=region_b),
        "subtitle": T["subtitle_region"].format(year=year_label),
        "hero_rates": build_hero_rates(blocks_b, lang),
        "intro": {"number": 1, "title": C["intro_title"],
                  "text": T["intro_region"].format(a=region_a, b=region_b, year=year_label)},
        "glossary": build_glossary(lang, number=2),
        "distribution": build_distribution(blocks_b, lang, number=3),
        "compare_bars": {"number": 4, "title": T["compare_section_title"],
                          "entities": [region_a, region_b], "items": items},
        "key_points": build_key_points(lang, number=5),
        "source": T["source"],
    }


def build_quarter_compare_spec(blocks_list, quarter_labels, year_label, lang):
    T = COMPARE_I18N[lang]
    C = DEMO_I18N[lang]
    items = []
    for code, label in zip(["TA", "TE", "TC"], T["indicators"]):
        values, pcts = [], []
        for blocks_q in blocks_list:
            v, p = get_rate(blocks_q, code, lang=lang)
            values.append(v)
            pcts.append(p)
        items.append({"label": label, "values": values, "pcts": pcts})

    return {
        "kicker": C["kicker"],
        "title": T["title_quarter"].format(year=year_label),
        "subtitle": T["subtitle_quarter"],
        "hero_rates": build_hero_rates(blocks_list[-1], lang),
        "intro": {"number": 1, "title": C["intro_title"], "text": T["intro_quarter"].format(year=year_label)},
        "glossary": build_glossary(lang, number=2),
        "distribution": build_distribution(blocks_list[-1], lang, number=3),
        "compare_bars": {"number": 4, "title": T["compare_section_title"],
                          "entities": quarter_labels, "items": items},
        "key_points": build_key_points(lang, number=5),
        "source": T["source"],
    }


# ==========================================================================
# 6. CLI
# ==========================================================================

def main():
    p = argparse.ArgumentParser(description="Génère une affiche HCP multilingue (FR/AR/EN) — 4 types disponibles.")
    p.add_argument("excel", help="Fichier Excel principal (région A / année A / trimestre 1 selon le mode)")
    p.add_argument("--mode", choices=["standard", "year_compare", "region_compare", "quarter_compare"],
                    default="standard", help="Type d'affiche à générer")
    p.add_argument("--sheet", default="2025", help="Feuille/année du fichier principal")
    p.add_argument("--lang", choices=["fr", "ar", "en"], default="fr", help="Langue de l'affiche")
    p.add_argument("--list", action="store_true", help="Liste les tableaux disponibles (fichier+feuille principaux) et quitte")
    p.add_argument("--config", help="Fichier JSON décrivant l'affiche (mode standard uniquement)")
    p.add_argument("--demo", action="store_true", help="Utilise la configuration de démonstration (mode standard)")
    p.add_argument("--output", default="affiche.png", help="Fichier image de sortie")

    g_year = p.add_argument_group("--mode year_compare")
    g_year.add_argument("--sheet-b", help="Feuille/année B à comparer à --sheet (année A)")

    g_region = p.add_argument_group("--mode region_compare")
    g_region.add_argument("--excel-b", help="Second fichier Excel (région B)")
    g_region.add_argument("--region-a", default="Région A", help="Nom affiché de la région A (fichier principal)")
    g_region.add_argument("--region-b", default="Région B", help="Nom affiché de la région B (--excel-b)")
    g_region.add_argument("--year", help="Feuille/année commune aux 2 fichiers (défaut : --sheet)")

    g_quarter = p.add_argument_group("--mode quarter_compare")
    g_quarter.add_argument("--q2", help="Fichier Excel du 2e trimestre")
    g_quarter.add_argument("--q3", help="Fichier Excel du 3e trimestre")
    g_quarter.add_argument("--q4", help="Fichier Excel du 4e trimestre")

    args = p.parse_args()

    if args.list:
        blocks = parse_ene_excel(args.excel, args.sheet)
        for titre, bloc in blocks.items():
            print(f"- {titre}")
            print(f"    lignes: {list(bloc['rows'])}")
            print(f"    milieux: {bloc['milieux']}  |  unité: {bloc['unite']}")
        return

    if args.mode == "standard":
        blocks = parse_ene_excel(args.excel, args.sheet)
        if args.config:
            with open(args.config, encoding="utf-8") as f:
                raw_spec = json.load(f)
            spec = build_from_excel(blocks, raw_spec, args.lang)
        elif args.demo:
            spec = build_demo_spec(blocks, args.lang)
        else:
            p.error("Mode standard : fournissez --config fichier.json ou --demo.")

    elif args.mode == "year_compare":
        if not args.sheet_b:
            p.error("--mode year_compare nécessite --sheet-b (2e année à comparer)")
        from openpyxl import load_workbook as _lwb
        all_sheets = _lwb(args.excel, read_only=True).sheetnames
        y0, y1 = sorted([int(args.sheet), int(args.sheet_b)])
        candidate_years = [s for s in all_sheets if s.isdigit() and y0 <= int(s) <= y1]
        blocks_by_year = {}
        for yr in candidate_years:
            try:
                blocks_by_year[yr] = parse_ene_excel(args.excel, yr)
            except Exception:
                continue  # feuille présente mais format incompatible : on l'ignore
        if str(args.sheet) not in blocks_by_year or str(args.sheet_b) not in blocks_by_year:
            p.error("Les feuilles --sheet et --sheet-b doivent être présentes et au format compatible.")
        spec = build_year_compare_spec(blocks_by_year, args.sheet, args.sheet_b, args.lang)

    elif args.mode == "region_compare":
        if not args.excel_b:
            p.error("--mode region_compare nécessite --excel-b (2e fichier Excel)")
        year = args.year or args.sheet
        blocks_a = parse_ene_excel(args.excel, year)
        blocks_b = parse_ene_excel(args.excel_b, year)
        spec = build_region_compare_spec(blocks_a, blocks_b, args.region_a, args.region_b, year, args.lang)

    elif args.mode == "quarter_compare":
        year = args.year or args.sheet
        files = [args.excel]
        labels = ["T1"]
        for i, fpath in enumerate([args.q2, args.q3, args.q4], start=2):
            if fpath:
                files.append(fpath)
                labels.append(f"T{i}")
        if len(files) < 2:
            p.error("--mode quarter_compare nécessite au moins --q2 (2 trimestres minimum)")
        blocks_list = [parse_ene_excel(f, year) for f in files]
        spec = build_quarter_compare_spec(blocks_list, labels, year, args.lang)

    out = create_affiche(spec, args.lang, args.output)
    print(f"Affiche générée ({args.mode}, {args.lang}) : {out}")


if __name__ == "__main__":
    main()
