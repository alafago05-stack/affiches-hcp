#!/usr/bin/env python3
"""parsers/annual.py — Parseur du format Excel HCP "ENE-Indicateurs
désagrégés" : un classeur avec une feuille par année, chaque feuille
contenant une succession de blocs "titre / ligne des milieux / lignes de
données / métadonnées (Champ, Référence, Unité, Source)".

Ce module ne dessine rien : il transforme un fichier Excel en dictionnaire
Python ({titre_tableau: {...}}). Les modules `posters/*.py` s'en servent pour
construire le `spec` donné aux fonctions de rendu de `poster_engine.py`.

La logique de `parse_ene_excel`, `get_value`, `format_number`, `pct_value`,
`get_rate` et `_resolve_title` vient de `generate_affiche.py` (référence),
avec les mêmes TYPES d'exceptions natives (ValueError / KeyError) :
`build_year_compare_spec` s'appuie spécifiquement sur `except KeyError` pour
ignorer les années dont le format de feuille est incompatible — ne pas les
remplacer par une exception "maison" sous peine de casser ce repli.

Par rapport à la référence, le parseur tolère en plus les fichiers
"désordonnés" rencontrés en pratique :
  - lignes/colonnes vides insérées en haut ou à gauche de la feuille ;
  - espaces parasites (y compris insécables) dans les titres, libellés de
    lignes et de milieux ;
  - apostrophes typographiques (') à la place des droites (') dans les
    titres ;
  - nombres stockés comme texte ("41,3" ou "41.3" au lieu de 41.3) ;
  - libellés à la casse différente lors des recherches (get_value).
Et `validate_blocks` vérifie AVANT génération que toutes les données
requises par les affiches sont présentes, pour afficher à l'utilisateur la
liste exacte de ce qui manque plutôt qu'une erreur en cascade.
"""

import re

from openpyxl import load_workbook

# Feuilles techniques présentes dans le classeur mais qui ne sont pas des
# années de données (sommaire, légende, etc.) : à exclure de list_years().
_NON_YEAR_SHEETS = {"Avis aux utilisateurs", "Liste des tableaux", "Signes conventionnels"}


class AnnualExcelFormatError(Exception):
    """Levée quand le fichier fourni n'est pas un .xlsx lisible ou n'a pas la
    structure minimale attendue. Le message est pensé pour être affiché tel
    quel à l'utilisateur."""


# ---------------------------------------------------------------------------
# Normalisation des cellules (tolérance aux fichiers désordonnés)
# ---------------------------------------------------------------------------
def _clean_cell(v):
    """Espaces insécables → espaces, chaînes réduites/strippées ; une chaîne
    vide ou faite d'espaces devient None (cellule vide)."""
    if isinstance(v, str):
        s = v.replace("\xa0", " ").strip()
        return s if s else None
    return v


def _norm_text(s) -> str:
    """Forme canonique pour comparer des libellés : minuscules, apostrophes
    typographiques → droites, espaces multiples réduits."""
    s = str(s).replace("’", "'").replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip().lower()


def _coerce_number(v):
    """Convertit les nombres stockés comme texte ("41,3", "1 234", "41.3")
    en float ; laisse tout le reste inchangé."""
    if isinstance(v, str):
        t = v.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(t)
        except ValueError:
            return v
    return v


def _find_key(keys, wanted):
    """Cherche `wanted` parmi `keys` : correspondance exacte d'abord, puis
    insensible à la casse/espaces/apostrophes. None si introuvable."""
    if wanted in keys:
        return wanted
    wn = _norm_text(wanted)
    for k in keys:
        if _norm_text(k) == wn:
            return k
    return None


def _is_meta(text) -> bool:
    """Ligne de métadonnées de fin de bloc (Champ / Référence / Unité /
    Source), quelle que soit la casse."""
    return _norm_text(text).startswith(("champ", "réf", "unit", "source"))


def list_years(path_or_file) -> list[str]:
    """Retourne la liste des feuilles qui ressemblent à des années (ex.
    '2019'..'2025'), dans l'ordre où elles apparaissent dans le classeur."""
    try:
        wb = load_workbook(path_or_file, data_only=True, read_only=True)
    except Exception as exc:
        raise AnnualExcelFormatError(
            f"Impossible de lire ce fichier Excel : {exc}. "
            "Vérifiez qu'il s'agit bien d'un fichier .xlsx valide."
        ) from exc
    years = [s for s in wb.sheetnames if s not in _NON_YEAR_SHEETS and re.fullmatch(r"\d{4}", s.strip())]
    if not years:
        raise AnnualExcelFormatError(
            "Aucune feuille ressemblant à une année (ex. '2025') n'a été trouvée dans ce "
            f"classeur. Feuilles présentes : {wb.sheetnames}"
        )
    return years


def parse_ene_excel(path: str, sheet: str) -> dict:
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        raise AnnualExcelFormatError(
            f"Impossible de lire ce fichier Excel : {exc}. "
            "Vérifiez qu'il s'agit bien d'un fichier .xlsx valide."
        ) from exc
    if sheet not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet}' introuvable. Feuilles dispo: {wb.sheetnames}")
    ws = wb[sheet]
    rows = [tuple(_clean_cell(v) for v in r) for r in ws.iter_rows(values_only=True)]

    # Tolérance aux fichiers décalés : on retire les colonnes entièrement
    # vides à gauche de la feuille (les lignes vides en haut sont déjà
    # ignorées par la boucle principale).
    first_col = min(
        (next(c for c, v in enumerate(r) if v is not None) for r in rows if any(v is not None for v in r)),
        default=0,
    )
    if first_col:
        rows = [r[first_col:] for r in rows]

    blocks, i, n = {}, 0, len(rows)
    while i < n:
        row = rows[i]
        title = row[0] if row else None
        if title and i + 1 < n and len(rows[i + 1]) > 1 and rows[i + 1][1]:
            next_row = rows[i + 1]
            # Les feuilles plus anciennes (2019, 2020...) intercalent une
            # ligne d'années (ex. 2019 en colonne 1, 2020 en colonne 7) entre
            # le titre et la ligne des milieux : chaque année a son propre
            # groupe de colonnes (Urbain/Rural/Ensemble × Chiffre/Signe) dans
            # le MÊME bloc. Les feuilles récentes (2022-2025) n'ont qu'une
            # seule année par feuille et vont directement à la ligne des
            # milieux. On détecte ce cas et on ne garde que le groupe de
            # colonnes de l'année demandée (`sheet`).
            nx1 = _coerce_number(next_row[1])
            is_year_row = isinstance(nx1, (int, float)) and 1900 <= nx1 <= 2100
            if is_year_row and i + 2 < n:
                year_row = [_coerce_number(v) for v in next_row]
                milieu_row = rows[i + 2]
                year_starts = [(int(v), c) for c, v in enumerate(year_row) if isinstance(v, (int, float))]
                bounds = [c for _, c in year_starts] + [len(milieu_row)]
                target_col = next((c for y, c in year_starts if str(y) == str(sheet)), None)
                if target_col is None:
                    # L'année demandée n'apparaît pas dans ce bloc précis :
                    # on saute tout le bloc (titre + lignes + métadonnées).
                    i += 1
                    while i < n and rows[i] and rows[i][0] and not _is_meta(rows[i][0]):
                        i += 1
                    while i < n and rows[i] and rows[i][0] and _is_meta(rows[i][0]):
                        i += 1
                    continue
                col_end = bounds[bounds.index(target_col) + 1]
                milieux, col_idx = [], []
                for c in range(target_col, col_end, 2):
                    if c < len(milieu_row) and milieu_row[c]:
                        milieux.append(str(milieu_row[c]))
                        col_idx.append(c)
                i += 4  # titre, ligne années, ligne milieux, ligne Chiffre/Signe
            else:
                milieu_row = next_row
                milieux, col_idx = [], []
                for c in range(1, len(milieu_row), 2):
                    if milieu_row[c]:
                        milieux.append(str(milieu_row[c]))
                        col_idx.append(c)
                i += 3
            data_rows = {}
            while i < n and rows[i] and rows[i][0] and not _is_meta(rows[i][0]):
                r = rows[i]
                vals = {}
                for milieu, c in zip(milieux, col_idx):
                    valeur = _coerce_number(r[c]) if c < len(r) else None
                    signe = r[c + 1] if c + 1 < len(r) else None
                    vals[milieu] = {"valeur": valeur, "signe": signe}
                data_rows[str(r[0])] = vals
                i += 1
            meta = {"champ": "", "reference": "", "unite": "", "source": ""}
            while i < n and rows[i] and rows[i][0] and _is_meta(rows[i][0]):
                text = str(rows[i][0])
                key, _, val = text.partition(":")
                key = key.strip().lower()
                if key.startswith("champ"):
                    meta["champ"] = val.strip()
                elif key.startswith("réf"):
                    meta["reference"] = val.strip()
                elif key.startswith("unit"):
                    meta["unite"] = val.strip()
                elif key.startswith("source"):
                    meta["source"] = val.strip()
                i += 1
            if not meta["unite"]:
                # Les feuilles anciennes n'ont pas toujours de ligne
                # "Unité : ..." séparée — l'unité est alors seulement
                # indiquée dans le titre lui-même (ex. "(%)", "(en milliers)").
                title_l = str(title).lower()
                if "%" in title_l:
                    meta["unite"] = "%"
                elif "millier" in title_l:
                    meta["unite"] = "Millier"
            blocks[str(title)] = {"milieux": milieux, "rows": data_rows, **meta}
        else:
            i += 1
    return blocks


def get_value(blocks, titre, ligne="Ensemble", milieu="Ensemble"):
    titre_k = _find_key(blocks, titre)
    if titre_k is None:
        raise KeyError(
            f"Tableau introuvable : {titre!r}. Tableaux disponibles dans cette feuille : {list(blocks)[:8]}"
        )
    bloc = blocks[titre_k]
    ligne_k = _find_key(bloc["rows"], ligne)
    if ligne_k is None:
        raise KeyError(
            f"Ligne {ligne!r} introuvable dans le tableau {titre_k!r}. "
            f"Lignes disponibles : {list(bloc['rows'])}"
        )
    milieu_k = _find_key(bloc["rows"][ligne_k], milieu)
    if milieu_k is None:
        raise KeyError(
            f"Milieu {milieu!r} introuvable dans le tableau {titre_k!r}. "
            f"Milieux disponibles : {bloc['milieux']}"
        )
    return bloc["rows"][ligne_k][milieu_k]["valeur"], bloc["unite"]


def format_number(valeur, unite, lang="fr"):
    if valeur is None:
        return {"fr": "n.d.", "en": "n/a", "ar": "غ.م"}[lang]
    sep = "," if lang != "en" else "."
    if "%" in (unite or ""):
        return f"{valeur:.1f}".replace(".", sep) + "%"
    if "millier" in (unite or "").lower():
        return f"{int(round(valeur * 1000)):,}".replace(",", " ")
    if isinstance(valeur, float) and valeur.is_integer():
        return f"{int(valeur):,}".replace(",", " ")
    return str(valeur)


def pct_value(valeur, unite):
    if valeur is None or "%" not in (unite or ""):
        return None
    return float(valeur)


# Table des codes internes utilisés par les posters vers les intitulés exacts
# des tableaux dans le classeur Excel.
INDICATOR_TITLES = {
    "TA": "Le taux d'activité par sexe et par milieu",
    "TE": "Le taux d'emploi par sexe et par milieu",
    "TC": "Le taux de chômage par sexe et par milieu",
    "TS": "Le taux de sous-emploi par sexe et par milieu",
}

FALLBACK_KEYWORDS = {
    "TA": ["taux d'activité"],
    "TE": ["taux d'emploi"],
    "TC": ["taux de chômage"],
    "TS": ["taux de sous-emploi", "sous-emploi"],
}

_INDICATOR_LABELS_FR = {
    "TA": "taux d'activité",
    "TE": "taux d'emploi",
    "TC": "taux de chômage",
    "TS": "taux de sous-emploi",
}


def _resolve_title(blocks, code):
    """Le libellé exact des tableaux varie parfois d'une feuille/année à
    l'autre dans les fichiers HCP (ex: 'Le taux d'activité (%)' en 2019-2020
    vs 'Le taux d'activité par sexe et par milieu' en 2022-2025). On essaie
    d'abord le titre exact connu, puis on cherche par mot-clé (comparaison
    insensible à la casse, aux espaces et au type d'apostrophe)."""
    titre = _find_key(blocks, INDICATOR_TITLES[code])
    if titre is not None:
        return titre
    for kw in FALLBACK_KEYWORDS.get(code, []):
        kw_n = _norm_text(kw)
        for t in blocks:
            if kw_n in _norm_text(t):
                return t
    raise KeyError(
        f"Aucun tableau de {_INDICATOR_LABELS_FR.get(code, code)} n'a été trouvé dans cette "
        "feuille — vérifiez que le fichier Excel contient bien ce tableau."
    )


def get_rate(blocks, code, milieu="Ensemble", ligne="Ensemble", lang="fr"):
    """Raccourci pour récupérer un indicateur (TA/TE/TC/TS) déjà formaté +
    sa valeur en % brute. Utilisé par les 4 types d'affiche."""
    titre = _resolve_title(blocks, code)
    valeur, unite = get_value(blocks, titre, ligne, milieu)
    return format_number(valeur, unite, lang), pct_value(valeur, unite)


# Même repli que _resolve_title, pour le tableau utilisé par
# `poster_engine.build_distribution` (répartition de la population). Le
# titre exact diffère lui aussi entre feuilles anciennes et récentes (ex.
# "Les 15 ans et plus par type d'activité (3) (en milliers)" en 2019-2020
# vs "La population en âge de travailler par type d'activité (3) et par
# milieu" en 2022-2025).
WORKING_AGE_TITLE = "La population en âge de travailler par type d'activité (3) et par milieu"
_WORKING_AGE_KEYWORDS = ["15 ans et plus", "âge de travailler", "âge d'activité"]


# Tableau du statut professionnel (3) des travailleurs — sert au taux de
# salariat (ligne « Salarié ») du nouveau design. Le titre varie lui aussi :
# "Le statut professionnel (3) des travailleurs (%)" en 2019-2020 vs
# "... par milieu" en 2021-2025. Attention aux faux amis : "Le statut
# professionnel recherché (3) par les chômeurs" et "Le taux de sous-emploi
# par statut professionnel (3)" — d'où l'ancrage sur "des travailleurs".
def resolve_salariat_title(blocks):
    for t in blocks:
        if "statut professionnel (3) des travailleurs" in _norm_text(t):
            return t
    raise KeyError(
        "Aucun tableau « statut professionnel (3) des travailleurs » (taux de "
        "salariat) n'a été trouvé dans cette feuille."
    )


def resolve_working_age_title(blocks):
    titre = _find_key(blocks, WORKING_AGE_TITLE)
    if titre is not None:
        return titre
    for t in blocks:
        tl = _norm_text(t)
        if "type d'activité" in tl and any(_norm_text(kw) in tl for kw in _WORKING_AGE_KEYWORDS):
            return t
    raise KeyError(
        "Aucun tableau de population en âge de travailler (ou '15 ans et plus par type "
        "d'activité') n'a été trouvé dans cette feuille."
    )


# ---------------------------------------------------------------------------
# Validation avant génération
# ---------------------------------------------------------------------------
def validate_blocks(blocks) -> list[str]:
    """Vérifie que la feuille contient toutes les données dont les affiches
    ont besoin. Retourne la liste des éléments manquants, en français lisible
    (liste vide = tout est là). Ne lève jamais d'exception.

    Couvre exactement ce que consomment poster_engine et posters/*.py :
      - TA/TE/TC ligne « Ensemble » avec une VALEUR numérique (anneaux héros
        et diagramme de répartition) ;
      - TA/TE/TC ligne « Féminin » et TS « Ensemble » (les lignes doivent
        exister ; une valeur vide s'affiche « n.d. » et n'est donc pas
        bloquante) ;
      - tableau de population en âge de travailler avec les lignes Ensemble /
        Actif occupé / Chômeur / Inactif et des valeurs numériques (le
        diagramme de flux fait des calculs dessus).
    """
    if not blocks:
        return [
            "aucun tableau ENE reconnu dans cette feuille — la structure ne "
            "correspond pas au format « ENE-Indicateurs désagrégés » attendu"
        ]
    missing = []
    # Les 3 taux essentiels : la valeur « Ensemble » doit exister ET être un nombre.
    nb_taux_reconnus = 0
    for code in ("TA", "TE", "TC"):
        label = _INDICATOR_LABELS_FR[code]
        try:
            titre = _resolve_title(blocks, code)
            nb_taux_reconnus += 1
            valeur, _ = get_value(blocks, titre)
            if not isinstance(valeur, (int, float)):
                missing.append(f"le {label} (valeur « Ensemble » vide ou non numérique)")
            elif not 0 <= valeur <= 100:
                # Un taux hors 0-100 % est forcément une erreur de saisie :
                # on refuse de générer une affiche avec un chiffre absurde.
                missing.append(
                    f"le {label} a une valeur suspecte ({valeur:g} %) — un taux doit être compris entre 0 et 100 %"
                )
        except KeyError:
            missing.append(f"le {label} (tableau ou ligne « Ensemble » introuvable)")
    if nb_taux_reconnus == 0:
        # Aucun des trois taux n'existe : ce n'est visiblement pas un export
        # ENE — un message court est plus clair que la liste exhaustive.
        return [
            "aucun tableau ENE reconnu dans cette feuille — la structure ne "
            "correspond pas au format « ENE-Indicateurs désagrégés » attendu"
        ]
    # Lignes secondaires : elles doivent exister (valeur vide tolérée → « n.d. »).
    for code in ("TA", "TE", "TC"):
        label = _INDICATOR_LABELS_FR[code]
        try:
            titre = _resolve_title(blocks, code)
            get_value(blocks, titre, "Féminin")
        except KeyError:
            missing.append(f"la ligne « Féminin » du {label}")
    try:
        titre = _resolve_title(blocks, "TS")
        get_value(blocks, titre)
    except KeyError:
        missing.append("le taux de sous-emploi (ligne « Ensemble »)")
    # NB : le taux de salariat (statut professionnel (3) des travailleurs)
    # n'est volontairement PAS bloquant — build_indicators affiche « n.d. »
    # si le tableau manque, l'affiche reste générable.
    # Répartition de la population : 4 lignes avec valeurs numériques.
    try:
        titre_pop = resolve_working_age_title(blocks)
        for ligne in ("Ensemble", "Actif occupé", "Chômeur", "Inactif"):
            try:
                valeur, _ = get_value(blocks, titre_pop, ligne)
                if not isinstance(valeur, (int, float)):
                    missing.append(f"la valeur « {ligne} » de la population en âge de travailler (vide ou non numérique)")
                elif valeur < 0:
                    missing.append(
                        f"la valeur « {ligne} » de la population en âge de travailler est négative ({valeur:g})"
                    )
            except KeyError:
                missing.append(f"la ligne « {ligne} » de la population en âge de travailler")
    except KeyError:
        missing.append("le tableau de la population en âge de travailler (15 ans et plus par type d'activité)")
    return missing


def guess_region_from_filename(filename: str) -> str:
    """Déduit un nom de région à partir du nom de fichier Excel (ex.
    'ENE-Indicateurs désagrégés 2019-2025- Région #09 -07052025.xlsx' ->
    'Région #09'). Repli sur le nom de fichier sans extension si aucun motif
    n'est reconnu."""
    stem = filename.rsplit(".", 1)[0]
    m = re.search(r"[Rr]égion\s*#?\s*\d+|[Rr]égion\s+[\wÀ-ÿ\- ]+", stem)
    if m:
        return m.group(0).strip(" -")
    return stem.strip(" -")
